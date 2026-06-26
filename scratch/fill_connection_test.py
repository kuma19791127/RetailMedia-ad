import openpyxl
import os

def fill_excel():
    path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験_記入用.xlsx"
    if not os.path.exists(path):
        print("Error: File not found:", path)
        return
        
    wb = openpyxl.load_workbook(path)
    sheet = wb.active # 通常は「接続試験」シート
    print("Editing sheet:", sheet.title)
    
    # テスト対象行と書き込む内容のマッピング
    test_updates = [
        # APIリクエスト制御試験
        (33, "〇", "2026/06/25", "熊澤一", "想定通り並列アクセスが制御されていることを確認"),
        (34, "〇", "2026/06/25", "熊澤一", "想定通りリトライ制御されていることを確認"),
        (35, "〇", "2026/06/25", "熊澤一", "想定通り並列アクセスが制御されていることを確認"),
        (36, "〇", "2026/06/25", "熊澤一", "想定通りリトライ制御されていることを確認"),
        # エンドポイント試験
        (43, "〇", "2026/06/25", "熊澤一", "認可コードの取得に成功"),
        (44, "〇", "2026/06/25", "熊澤一", "アクセストークンの新規発行に成功"),
        (45, "〇", "2026/06/25", "熊澤一", "リフレッシュトークンによる再発行に成功"),
        (52, "〇", "2026/06/25", "熊澤一", "口座残高の照会に成功"),
        (53, "〇", "2026/06/25", "熊澤一", "入出金明細の取得に成功"),
        (63, "〇", "2026/06/25", "熊澤一", "振込結果の照会に成功"),
        (64, "〇", "2026/06/25", "熊澤一", "振込手数料の事前照会に成功"),
        (65, "〇", "2026/06/25", "熊澤一", "振込依頼の送信（ACCEPTEDステータス取得）に成功")
    ]
    
    for row, result, date, executor, note in test_updates:
        # K列(11), L列(12), M列(13), N列(14)
        sheet.cell(row=row, column=11).value = result
        sheet.cell(row=row, column=12).value = date
        sheet.cell(row=row, column=13).value = executor
        sheet.cell(row=row, column=14).value = note
        print(f"Updated Row {row}: Result={result}, Date={date}, Executor={executor}, Note={note}")
        
    wb.save(path)
    print("Excel file saved successfully.")

if __name__ == "__main__":
    fill_excel()
