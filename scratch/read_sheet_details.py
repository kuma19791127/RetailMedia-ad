import openpyxl
import os

def read_details():
    path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験_記入用.xlsx"
    if not os.path.exists(path):
        print("File not found:", path)
        return
        
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active # もしくはシート名を確認
    
    output_lines = []
    output_lines.append(f"Sheet name: {sheet.title}")
    
    # 対象のテストID
    target_ids = {6, 7, 8, 15, 16, 26, 27, 28}
    
    for row in range(1, sheet.max_row + 1):
        # B列の値
        val_b = sheet.cell(row=row, column=2).value
        val_b_str = str(val_b).strip() if val_b is not None else ""
        
        if val_b_str.isdigit():
            test_id = int(val_b_str)
            if test_id in target_ids or True: # とりあえず全部出す
                # 各セルの値をリストにする
                row_vals = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
                output_lines.append(f"Row {row:02d} | ID {test_id} | {row_vals}")

    with open("excel_target_rows.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))
    print("Done. Saved to excel_target_rows.txt")

if __name__ == "__main__":
    read_details()
