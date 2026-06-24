import openpyxl
import os

def verify():
    path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験_記入用.xlsx"
    if not os.path.exists(path):
        print("File not found")
        return
    
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["接続試験"]
    
    print("--- Row 28 & 29 (Dates) ---")
    print(f"Row 28: B={sheet.cell(28,2).value}, D={sheet.cell(28,4).value}")
    print(f"Row 29: B={sheet.cell(29,2).value}, D={sheet.cell(29,4).value}")
    
    print("\n--- Row 33 (Request Limits) ---")
    print(f"Row 33: B={sheet.cell(33,2).value}, Scope={sheet.cell(33,6).value}, API={sheet.cell(33,8).value}")
    
    print("\n--- Endpoint tests samples ---")
    for r in [39, 42, 43, 52, 53, 54, 63, 64, 65]:
        print(f"Row {r}: TestID={sheet.cell(r,2).value}, API={sheet.cell(r,8).value}, Result={sheet.cell(r,11).value}, Note={sheet.cell(r,14).value}")

if __name__ == "__main__":
    verify()
