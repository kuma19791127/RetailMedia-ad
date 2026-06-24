import openpyxl
import os
import re

files_to_check = [
    r"C:\Users\one\Desktop\API利用申込書.xlsx",
    r"C:\Users\one\Desktop\コピーAPI連携ヒアリングシート兼チェックリストnon-logi株式会社（更新系）.xlsx"
]

ip_pattern = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")

for filepath in files_to_check:
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        continue
    print(f"\n=========================================\nChecking Excel file: {filepath}")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                # Join row values as string to search for IP or interesting words
                row_str = " | ".join([str(val) for val in row if val is not None])
                if not row_str.strip():
                    continue
                # Search for IP, "IP", "アドレス", "届け出", "届出"
                if any(x in row_str.lower() for x in ["ip", "アドレス", "届け出", "届出", "接続", "gmo"]):
                    print(f"Row {row_idx:03d}: {row_str}")
                elif ip_pattern.search(row_str):
                    print(f"Row {row_idx:03d} (IP found): {row_str}")
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
