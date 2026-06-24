import openpyxl
import os

def edit_excel():
    src_path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験.xlsx"
    dest_path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験_記入用.xlsx"
    
    if not os.path.exists(src_path):
        print(f"Error: Source file {src_path} does not exist.")
        return
        
    wb = openpyxl.load_workbook(src_path)
    sheet = wb["接続試験"]
    
    # セクション位置を把握する
    in_request_control = False
    in_endpoint_test = False
    
    # 対象とするテストID
    target_ids = {6, 7, 8, 15, 16, 26, 27, 28}
    
    for row in range(1, sheet.max_row + 1):
        # B列の値
        val_b = sheet.cell(row=row, column=2).value
        # 文字列化して空白除去
        val_b_str = str(val_b).strip() if val_b is not None else ""
        
        # セクション切り替えの判定
        if "【APIリクエスト制御試験】" in val_b_str or "APIリクエスト制御試験" in str(sheet.cell(row=row, column=1).value):
            in_request_control = True
            in_endpoint_test = False
            continue
        elif "【エンドポイント試験】" in val_b_str or "エンドポイント試験" in str(sheet.cell(row=row, column=1).value):
            in_request_control = False
            in_endpoint_test = True
            continue
            
        # 1. 実施期間の設定
        if val_b_str == "試験実施開始日":
            print(f"Row {row}: Setting start date to 2026/06/25")
            sheet.cell(row=row, column=4).value = "2026/06/25"
        elif val_b_str == "試験実施終了日":
            print(f"Row {row}: Setting end date to 2026/06/25")
            sheet.cell(row=row, column=4).value = "2026/06/25"
            
        # 2. APIリクエスト制御試験セクションの処理
        if in_request_control and not in_endpoint_test:
            # F列 (col 6) の値をチェック
            val_f = sheet.cell(row=row, column=6).value
            if val_f == "ご利用するスコープを設定":
                sheet.cell(row=row, column=6).value = "private:account, private:transfer"
                print(f"Row {row}: Updated Scope")
            # H列 (col 8) の値をチェック
            val_h = sheet.cell(row=row, column=8).value
            if val_h == "ご利用するAPIを設定":
                sheet.cell(row=row, column=8).value = "残高照会, 振込依頼"
                print(f"Row {row}: Updated API")
                
        # 3. エンドポイント試験セクションの処理
        if in_endpoint_test:
            # B列がテストID (数値) であるか判定
            if val_b_str.isdigit():
                test_id = int(val_b_str)
                if test_id not in target_ids:
                    # 対象外を設定
                    sheet.cell(row=row, column=11).value = "対象外" # K列 (実行結果)
                    sheet.cell(row=row, column=14).value = "リテアドでは使用しないため対象外" # N列 (備考)
                    # print(f"Row {row}: Set TestID {test_id} to TargetExempt")
                else:
                    # 対象項目は空欄にする（テスト当日に記入するためクリア）
                    sheet.cell(row=row, column=11).value = ""
                    sheet.cell(row=row, column=14).value = ""
                    print(f"Row {row}: Set TestID {test_id} as Target (Cleared fields)")

    wb.save(dest_path)
    print(f"Saved modified Excel to: {dest_path}")

if __name__ == "__main__":
    edit_excel()
