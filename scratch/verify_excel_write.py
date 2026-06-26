import openpyxl
import os

def verify():
    path = r"C:\Users\one\Desktop\コピー【オープンAPI】事業者接続試験_記入用.xlsx"
    if not os.path.exists(path):
        print("File not found:", path)
        return
        
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    
    rows_to_check = [33, 34, 35, 36, 43, 44, 45, 52, 53, 63, 64, 65]
    print(f"=== Verifying sheet: {sheet.title} ===")
    for row in rows_to_check:
        k = sheet.cell(row=row, column=11).value
        l = sheet.cell(row=row, column=12).value
        m = sheet.cell(row=row, column=13).value
        n = sheet.cell(row=row, column=14).value
        # 日本語を含むので、UTF-8エンコードされたテキストファイルへログ出力して確認する
        print(f"Row {row}: K={k}, L={l}, M={m}, N={n}")
        
    with open("verify_result.txt", "w", encoding="utf-8") as f:
        f.write(f"=== Verifying sheet: {sheet.title} ===\n")
        for row in rows_to_check:
            k = sheet.cell(row=row, column=11).value
            l = sheet.cell(row=row, column=12).value
            m = sheet.cell(row=row, column=13).value
            n = sheet.cell(row=row, column=14).value
            f.write(f"Row {row}: K={k}, L={l}, M={m}, N={n}\n")
            
if __name__ == "__main__":
    verify()
