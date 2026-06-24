import openpyxl
import os

def read_excel():
    path = "Book2.xlsx"
    if not os.path.exists(path):
        print(f"File {path} not found.")
        return
        
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Loaded {path}. Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        print(f"\n--- Sheet: {name} ---")
        sheet = wb[name]
        for row in sheet.iter_rows(values_only=True):
            if any(row):  # Row is not empty
                print(row)

if __name__ == "__main__":
    read_excel()
