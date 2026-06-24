import openpyxl
import os
import glob
import sys

def dump_excel_content():
    desktop_path = r"C:\Users\one\Desktop"
    xlsx_files = glob.glob(os.path.join(desktop_path, "*.xlsx"))
    
    # ワークスペース内の test.xlsx も含める
    workspace_path = r"c:\Users\one\.gemini\antigravity\playground\twilight-parsec"
    workspace_xlsx = glob.glob(os.path.join(workspace_path, "*.xlsx"))
    xlsx_files.extend(workspace_xlsx)

    output_path = os.path.join(workspace_path, "scratch", "excel_dump.txt")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=== EXCEL FILES DUMP ===\n\n")
        for filepath in xlsx_files:
            basename = os.path.basename(filepath)
            size = os.path.getsize(filepath)
            f.write(f"=========================================\n")
            f.write(f"FILE: {basename} ({size} bytes)\n")
            f.write(f"PATH: {filepath}\n")
            f.write(f"=========================================\n")
            
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                f.write(f"Sheets: {wb.sheetnames}\n\n")
                for sheetname in wb.sheetnames:
                    f.write(f"--- Sheet: {sheetname} ---\n")
                    sheet = wb[sheetname]
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True):
                        if any(row is not None and str(row).strip() != "" for row in row):
                            # 空白行でなければ出力するが、大きすぎる場合は制限する
                            row_count += 1
                            if row_count > 150:
                                f.write("... (truncated after 150 rows) ...\n")
                                break
                            # Noneを空文字にして出力
                            row_str = ", ".join([str(val) if val is not None else "" for val in row])
                            f.write(f"Row {row_count}: {row_str}\n")
                    f.write("\n")
            except Exception as e:
                f.write(f"Error reading file: {e}\n\n")
            f.write("\n")
            
    print(f"Dumped excel content to {output_path}")

if __name__ == "__main__":
    dump_excel_content()
